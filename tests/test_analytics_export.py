"""Юнит-тесты выгрузки аналитики в .xlsx/.pdf (п.5–6 баг-репорта 06.08.2026).

Проверяем ровно то, на что жаловался заказчик: дата в ячейке должна быть ДАТОЙ
(а не строкой `2026-08-06T04:00:00.123456+00:00`), доля — процентом, счётчик —
целым, заголовки — выделенными, «весь отчёт» — одним файлом с листом на раздел.

БД и сеть не нужны: обе функции чистые, на вход подаётся тот же JSON, который
роутер получает из `crud.build_report(...).model_dump(mode="json", by_alias=True)`.

Запуск:
    python -m pytest tests/test_analytics_export.py
    python tests/test_analytics_export.py          # standalone

PDF-тесты требуют TTF с кириллицей. В образе он есть (fonts-dejavu-core), на
машине разработчика может не быть — такие тесты помечаются skip, а не падают.
"""
from __future__ import annotations

import contextlib
import os
import sys
import tempfile
from datetime import date, datetime
from io import BytesIO

import pytest

sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

from app.services import analytics_export as ax

openpyxl = pytest.importorskip("openpyxl")
from openpyxl import load_workbook  # noqa: E402  — только после importorskip


# ── Фикстуры-пейлоады ────────────────────────────────────────────────────────
# `updated_at` намеренно взят с микросекундами и смещением +03:00 — это дословно
# то, что приезжало в ячейку строкой (п.5). В UTC это 06.08.2026 04:00.
OVERVIEW = {
    "from": "2026-07-01",
    "to": "2026-08-06",
    "updated_at": "2026-08-06T07:00:00.123456+03:00",
    "total_sessions": 128,
    "total_visits": 141,
    "total_app_opens": 96,
    "total_recognitions": 302,
    "recognition_success_rate": 0.8421,
    "total_chat_messages": 57,
    "total_audio_plays": 40,
    "top_exhibits": [
        {"id": 7, "name": "Пасхальное яйцо «Ренессанс»", "count": 31},
        {"id": 531, "name": "Пасхальное яйцо-шкатулка «Ренессанс»", "count": 12},
    ],
    "top_halls": [{"id": 5, "name": "Золотая гостиная", "count": 74}],
}

# Длинный вопрос — под потолок ширины колонки; угловые скобки и амперсанд —
# под мини-XML reportlab (без экранирования сборка PDF падала целиком).
LONG_QUESTION = (
    "Расскажите, пожалуйста, подробно, как именно мастера фирмы Фаберже "
    "изготавливали пасхальные яйца с сюрпризом, сколько времени уходило на одно "
    "изделие и сколько человек над ним работало, и правда ли, что <мастер> "
    "Перхин & Вигстрём делали их вручную?"
)
QUESTIONS = {
    "from": "2026-07-01",
    "to": "2026-08-06",
    "updated_at": "2026-08-06T04:00:00+00:00",
    "total_questions": 214,
    "unique_questions": 173,
    "total_clusters": 96,
    "frequent": [
        {"question": LONG_QUESTION, "count": 9, "variants": ["как делали яйца", "a < b & c"]},
        {"question": "Сколько стоит яйцо?", "count": 7, "variants": []},
    ],
    "rare": [{"question": "Где купить сувенир?", "count": 1, "variants": []}],
}

# Цифры подобраны так, чтобы доля НЕ сходилась ни с одним «очевидным» частным:
# 12/96 = 12,5 %, при этом визитов 141, а чат открывали в 20 визитах за период.
# Ровно на этом и ловится дефект «на листе нечем проверить процент» (п.7).
ENGAGEMENT = {
    "from": "2026-07-01",
    "to": "2026-08-06",
    "updated_at": "2026-08-06T04:00:00+00:00",
    "total_visits": 141,
    "avg_duration_sec": 812.5,
    "median_duration_sec": 640.0,
    "sessions_with_chat": 20,
    "sessions_with_questions": 14,
    "sessions_with_app_open": 96,
    "conversion_basis": "app_open",
    "conversion_denominator": 96,
    "chat_conversion_numerator": 12,
    "question_conversion_numerator": 6,
    "chat_conversion_rate": 0.125,
    "question_conversion_rate": 0.0625,
    "buckets": [{"label": "5–15 мин", "count": 63}, {"label": "> 60 мин", "count": 4}],
}


def _sheet(report: str, payload: dict):
    """xlsx в память → активный лист (файл на диск не пишем — тестам он не нужен)."""
    return load_workbook(BytesIO(ax.to_xlsx(report, payload))).active


def _find_row(sheet, label: str) -> int:
    """Номер строки, у которой в колонке A стоит `label`."""
    for row in range(1, sheet.max_row + 1):
        if sheet.cell(row=row, column=1).value == label:
            return row
    raise AssertionError(f"строка «{label}» не найдена на листе")


def _require_font() -> None:
    """PDF без кириллического TTF не собрать — это не провал теста, а среда."""
    if not ax.font_available():
        pytest.skip("нет TTF с кириллицей: PDF-тесты пропущены (в образе шрифт есть)")


# ── Контракт с роутером ──────────────────────────────────────────────────────
def test_all_reports_is_six_dashboard_reports_plus_engagement():
    """п.6: в «весь отчёт» входят шесть плиток дашборда и метрики визита (п.7)."""
    assert ax.REPORTS == ("overview", "questions", "unanswered", "exhibits", "routes", "recognition")
    assert ax.ALL_REPORTS == ax.REPORTS + ("engagement",)


def test_file_name_for_all_is_analytics():
    assert ax.file_name("all", date(2026, 7, 1), date(2026, 8, 6), "xlsx") == (
        "faberge-analytics-2026-07-01-2026-08-06.xlsx"
    )
    assert ax.file_name("overview", None, None, "pdf") == "faberge-overview-all-all.pdf"


# ── п.5: типы ячеек ──────────────────────────────────────────────────────────
def test_xlsx_opens_as_workbook():
    data = ax.to_xlsx("overview", OVERVIEW)
    assert data[:2] == b"PK"                       # xlsx — это zip
    assert load_workbook(BytesIO(data)).active.title == "Сводная аналитика"


def test_updated_at_is_datetime_cell_without_microseconds():
    """Главная претензия п.5: «дата и время с ненужными данными»."""
    sheet = _sheet("overview", OVERVIEW)
    cell = sheet.cell(row=_find_row(sheet, "Данные на"), column=2)
    assert isinstance(cell.value, datetime), "в ячейке строка, а не дата"
    assert cell.value == datetime(2026, 8, 6, 4, 0)   # +03:00 приведено к UTC
    assert cell.value.microsecond == 0
    assert cell.number_format == "DD.MM.YYYY HH:MM"
    assert cell.value.tzinfo is None                 # Excel не хранит таймзону


def test_period_cells_are_dates():
    sheet = _sheet("overview", OVERVIEW)
    row = _find_row(sheet, "Период")
    start, end = sheet.cell(row=row, column=2), sheet.cell(row=row, column=3)
    assert (start.value, end.value) == (datetime(2026, 7, 1), datetime(2026, 8, 6))
    assert start.number_format == end.number_format == "DD.MM.YYYY"


def test_generated_at_in_header():
    """Шапка листа: название, период, дата формирования файла."""
    sheet = _sheet("overview", OVERVIEW)
    assert sheet.cell(row=1, column=1).value == "Сводная аналитика"
    created = sheet.cell(row=_find_row(sheet, "Файл создан"), column=2)
    assert isinstance(created.value, datetime) and created.number_format == "DD.MM.YYYY HH:MM"


def test_rate_is_number_with_percent_format():
    """Доли в payload — float 0..1, в Excel это процентный формат, а не текст «84,2 %»."""
    sheet = _sheet("overview", OVERVIEW)
    cell = sheet.cell(row=_find_row(sheet, "Доля успешных распознаваний"), column=2)
    assert cell.value == 0.8421
    assert isinstance(cell.value, float)
    assert cell.number_format == "0.0%"


def test_counters_are_int():
    """«Числа суммируются формулами» из DoD — значит int, а не строка."""
    sheet = _sheet("overview", OVERVIEW)
    cell = sheet.cell(row=_find_row(sheet, "Сессий"), column=2)
    assert cell.value == 128 and isinstance(cell.value, int)
    assert cell.number_format == "0"
    assert cell.alignment.horizontal == "right"


def test_float_metric_stays_float():
    sheet = _sheet("engagement", ENGAGEMENT)
    cell = sheet.cell(row=_find_row(sheet, "Средняя длительность визита, с"), column=2)
    assert cell.value == 812.5 and isinstance(cell.value, float)
    assert cell.number_format == "0.##"


def test_conversion_denominator_is_exported():
    """п.7: знаменатель конверсий музей должен видеть рядом с самой конверсией."""
    sheet = _sheet("engagement", ENGAGEMENT)
    assert sheet.cell(row=_find_row(sheet, "Визитов с открытием приложения"), column=2).value == 96
    assert sheet.cell(row=_find_row(sheet, "Конверсия в чат"), column=2).number_format == "0.0%"


def test_conversion_ratio_can_be_checked_on_the_sheet():
    """Дробь должна сходиться прямо в файле: числитель / знаменатель = напечатанная доля.

    Раньше на лист уезжали только «Визитов», «Визитов с открытием чата» и сама доля — и ни
    20/141, ни 20/96 не равнялись 12,5 %, потому что числитель считается ВНУТРИ базы. Музей
    видел процент, который нечем проверить (п.7).
    """
    sheet = _sheet("engagement", ENGAGEMENT)
    denominator = sheet.cell(row=_find_row(sheet, "Знаменатель конверсий (визитов)"), column=2).value
    numerator = sheet.cell(row=_find_row(sheet, "Из них открыли чат"), column=2).value
    rate = sheet.cell(row=_find_row(sheet, "Конверсия в чат"), column=2).value
    assert (denominator, numerator) == (96, 12)
    assert abs(numerator / denominator - rate) < 1e-9


def test_conversion_basis_is_spelled_out_in_words():
    """`app_open` в отчёте для музея ничего не значит, а по базе решается сопоставимость периодов."""
    sheet = _sheet("engagement", ENGAGEMENT)
    value = sheet.cell(row=_find_row(sheet, "База конверсий"), column=2).value
    assert "app_open" in value and "запуск" in value
    sheet = _sheet("engagement", dict(ENGAGEMENT, conversion_basis="all_visits"))
    assert "все визиты" in sheet.cell(row=_find_row(sheet, "База конверсий"), column=2).value


# ── п.5: вёрстка листа ───────────────────────────────────────────────────────
def test_section_headers_are_bold_filled_and_frozen():
    sheet = _sheet("overview", OVERVIEW)
    head_row = _find_row(sheet, "Показатель")
    head = sheet.cell(row=head_row, column=1)
    assert head.font.bold is True
    assert head.fill.fgColor.rgb.endswith(ax._HEADER_FILL_COLOR)
    assert head.alignment.wrap_text is True
    assert head.border.bottom.style == "thin"
    # Закрепляем по первой секции: freeze_panes в Excel один на лист, а секций
    # на листе несколько — см. комментарий в _write_sheet.
    assert sheet.freeze_panes == f"A{head_row + 1}"


def test_second_section_headers_are_also_marked():
    """Секций на листе несколько — выделена шапка каждой, не только первой."""
    sheet = _sheet("overview", OVERVIEW)
    head = sheet.cell(row=_find_row(sheet, "ID"), column=1)
    assert head.font.bold is True and head.fill.fgColor.rgb.endswith(ax._HEADER_FILL_COLOR)


def test_long_text_wraps_and_column_width_is_capped():
    """Вопрос на 250 символов не должен растягивать колонку на два экрана."""
    sheet = _sheet("questions", QUESTIONS)
    row = _find_row(sheet, "Вопрос") + 1
    cell = sheet.cell(row=row, column=1)
    assert cell.value.startswith("Расскажите")
    assert cell.alignment.wrap_text is True
    assert cell.alignment.vertical == "top"
    width = sheet.column_dimensions["A"].width
    assert ax._MIN_COL_WIDTH <= width <= ax._MAX_COL_WIDTH


def test_text_is_left_aligned_numbers_right():
    sheet = _sheet("questions", QUESTIONS)
    row = _find_row(sheet, "Вопрос") + 1
    assert sheet.cell(row=row, column=1).alignment.horizontal == "left"
    assert sheet.cell(row=row, column=2).alignment.horizontal == "right"


# ── Распознавание ISO-строк ──────────────────────────────────────────────────
def test_iso_strings_become_dates():
    assert ax._as_cell("2026-08-06").fmt == "date"
    assert ax._as_cell("2026-08-06T04:00:00.123456+00:00").fmt == "datetime"
    assert ax._as_cell("2026-08-06T04:00:00Z").value == datetime(2026, 8, 6, 4, 0)


def test_plain_text_is_not_mistaken_for_a_date():
    """Названия экспонатов и вопросы датами становиться не должны."""
    for text in ("2026", "06.08.2026", "Яйцо 1900-х годов", "2026-13-45", "витрина 1-2-3"):
        assert ax._as_cell(text).fmt == "text", text


# ── п.6: весь отчёт одним файлом ─────────────────────────────────────────────
def test_xlsx_all_has_index_and_a_sheet_per_report():
    book = load_workbook(BytesIO(ax.to_xlsx_all({"overview": OVERVIEW, "engagement": ENGAGEMENT})))
    names = book.sheetnames
    assert names[0] == "Оглавление"
    assert len(names) == len(ax.ALL_REPORTS) + 1
    assert len(set(names)) == len(names)                       # имена уникальны
    for name in names:
        assert len(name) <= 31                                 # жёсткий лимит Excel
        assert not set(name) & set("[]:*?/\\")
    assert book["Сводная аналитика"].cell(row=1, column=1).value == "Сводная аналитика"


def test_xlsx_all_index_lists_sections_and_marks_empty_ones():
    book = load_workbook(BytesIO(ax.to_xlsx_all({"overview": OVERVIEW})))
    index = book["Оглавление"]
    values = [
        tuple(index.cell(row=row, column=col).value for col in (1, 3))
        for row in range(1, index.max_row + 1)
    ]
    assert index.cell(row=1, column=1).value == "Аналитика посетителей"
    assert ("Сводная аналитика", 10) in values                 # 7 метрик + 2 экспоната + 1 зал
    assert ("Метрики визита", "нет данных") in values          # раздел не потерялся


def test_sheet_name_is_trimmed_and_unique():
    used: set = set()
    long_title = "Очень длинное название отчёта про распознавание экспонатов"
    first = ax._sheet_name(long_title, used)
    second = ax._sheet_name(long_title, used)
    assert len(first) <= 31 and len(second) <= 31
    assert first != second
    assert ax._sheet_name("Топ: залы/витрины [2026]", used) == "Топ залы витрины 2026"


def test_empty_payload_does_not_break_xlsx():
    """DoD п.6: раздел без данных за период — с пометкой, а не с падением."""
    sheet = load_workbook(BytesIO(ax.to_xlsx("routes", {}))).active
    assert any(
        sheet.cell(row=row, column=1).value == "Нет данных за выбранный период"
        for row in range(1, sheet.max_row + 1)
    )
    book = load_workbook(BytesIO(ax.to_xlsx_all({})))
    assert len(book.sheetnames) == len(ax.ALL_REPORTS) + 1


# ── Текст из внешнего мира ───────────────────────────────────────────────────
def test_control_characters_and_huge_text_do_not_break_xlsx():
    """openpyxl падает целым файлом на \\x07 и на строке длиннее 32 767 символов."""
    payload = dict(QUESTIONS, frequent=[
        {"question": "вопрос\x07 с управляющим символом", "count": 1, "variants": ["ф" * 40000]},
    ])
    sheet = load_workbook(BytesIO(ax.to_xlsx("questions", payload))).active
    row = _find_row(sheet, "Вопрос") + 1
    assert "\x07" not in sheet.cell(row=row, column=1).value
    assert len(sheet.cell(row=row, column=3).value) <= 32767


# ── PDF (п.3 и п.6) ──────────────────────────────────────────────────────────
def test_font_available_returns_bool_and_never_raises():
    """/health зовёт это на каждом опросе — исключение оттуда недопустимо."""
    assert isinstance(ax.font_available(), bool)


def test_missing_font_gives_instructive_error():
    """Без шрифта — 503 с инструкцией «что доложить», а не лист с квадратами."""
    saved_candidates, saved_bundled = ax._FONT_CANDIDATES, ax._bundled_font_paths
    saved_path = ax.settings.analytics_pdf_font_path
    ax._FONT_CANDIDATES, ax._bundled_font_paths = (), (lambda: [])
    ax.settings.analytics_pdf_font_path = None
    try:
        assert ax.font_available() is False
        with pytest.raises(ax.ExportError) as info:
            ax._resolve_font_path()
        assert "ANALYTICS_PDF_FONT_PATH" in info.value.message
    finally:
        ax._FONT_CANDIDATES, ax._bundled_font_paths = saved_candidates, saved_bundled
        ax.settings.analytics_pdf_font_path = saved_path


def _forget_font() -> None:
    """Забыть зарегистрированный шрифт: reportlab держит его в глобальном кэше, и тест,
    запущенный ПОСЛЕ удачной сборки PDF, иначе даже не дошёл бы до чтения файла."""
    try:
        from reportlab.pdfbase import pdfmetrics
    except ImportError:                                # pragma: no cover — зависит от окружения
        return
    pdfmetrics._fonts.pop(ax._FONT_NAME, None)
    pdfmetrics._fonts.pop(ax._FONT_BOLD_NAME, None)


@contextlib.contextmanager
def _only_font(blob: bytes):
    """Единственный кандидат на шрифт — временный файл с этим содержимым."""
    path = os.path.join(tempfile.mkdtemp(), "font.ttf")
    with open(path, "wb") as fh:
        fh.write(blob)
    saved = (ax._FONT_CANDIDATES, ax._bundled_font_paths, ax.settings.analytics_pdf_font_path)
    ax._FONT_CANDIDATES, ax._bundled_font_paths = (), (lambda: [])
    ax.settings.analytics_pdf_font_path = path
    _forget_font()
    try:
        yield
    finally:
        ax._FONT_CANDIDATES, ax._bundled_font_paths, ax.settings.analytics_pdf_font_path = saved
        _forget_font()


def test_unusable_font_file_is_a_503_not_a_500():
    """Файл ЕСТЬ, но негодный: раньше /health рапортовал pdf_font=up, а выгрузка падала 500.

    Роутер ловит только ExportError, поэтому reportlab TTFError уходил наружу пустым 500 —
    ровно жалоба заказчика из п.3, но с ложным диагнозом в /health и без подсказки, что делать.
    Три случая проверены руками на живом reportlab: сохранённая страница прокси, пустой файл
    и OpenType/CFF (сигнатура OTTO — валидный шрифт, который ReportLab не читает).
    """
    cases = {
        "страница прокси вместо шрифта": b"<html><body>403 Forbidden</body></html>" * 8000,
        "пустой файл": b"",
        "OpenType/CFF": b"OTTO" + b"\x00" * 400_000,
    }
    for label, blob in cases.items():
        with _only_font(blob):
            assert ax.font_available() is False, label
            with pytest.raises(ax.ExportError) as info:
                ax.to_pdf("overview", OVERVIEW)
            assert "ANALYTICS_PDF_FONT_PATH" in info.value.message, label


def test_corrupted_font_with_valid_signature_still_gives_export_error():
    """Сигнатура sfnt на месте, а внутри мусор — последний рубеж внутри _register_fonts."""
    with _only_font(b"\x00\x01\x00\x00" + b"\x00" * 512):
        assert ax.font_available() is True             # сигнатуру такой файл проходит
        with pytest.raises(ax.ExportError) as info:
            ax.to_pdf("overview", OVERVIEW)
    assert "не читается ReportLab" in info.value.message


def test_pdf_is_generated():
    _require_font()
    data = ax.to_pdf("overview", OVERVIEW)
    assert data.startswith(b"%PDF") and len(data) > 1000


def test_pdf_survives_angle_brackets_in_a_question():
    """reportlab читает абзац как мини-XML: «<мастер>» ронял всю выгрузку."""
    _require_font()
    assert ax.to_pdf("questions", QUESTIONS).startswith(b"%PDF")


def test_pdf_survives_oversized_cell():
    """Строка выше страницы не разбивается reportlab'ом — текст обрезаем."""
    _require_font()
    payload = dict(QUESTIONS, frequent=[
        {"question": "вопрос", "count": 1, "variants": ["формулировка"] * 3000},
    ])
    assert ax.to_pdf("questions", payload).startswith(b"%PDF")


def test_pdf_all_is_one_document_with_every_report():
    _require_font()
    data = ax.to_pdf_all({"overview": OVERVIEW, "questions": QUESTIONS, "engagement": ENGAGEMENT})
    assert data.startswith(b"%PDF")
    # Разделы начинаются с новой страницы, поэтому страниц не меньше, чем разделов
    # (титул + по странице на отчёт).
    assert data.count(b"/Type /Page\n") >= len(ax.ALL_REPORTS)
    assert len(data) > len(ax.to_pdf("overview", OVERVIEW))


def test_empty_payload_does_not_break_pdf():
    _require_font()
    assert ax.to_pdf("routes", {}).startswith(b"%PDF")
    assert ax.to_pdf_all({}).startswith(b"%PDF")


if __name__ == "__main__":
    failures = 0
    for name, func in sorted(globals().items()):
        if not (name.startswith("test_") and callable(func)):
            continue
        try:
            func()
            print(f"ok   {name}")
        except AssertionError as exc:
            failures += 1
            print(f"FAIL {name}: {exc}")
        except pytest.skip.Exception as exc:           # нет шрифта — не провал
            print(f"skip {name}: {exc}")
    print("—" * 40)
    print("все тесты пройдены" if not failures else f"провалено: {failures}")
    sys.exit(1 if failures else 0)
