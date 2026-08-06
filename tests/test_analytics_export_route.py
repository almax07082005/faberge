"""Тесты ручки GET /admin/analytics/export (п.3 и п.6 баг-репорта 06.08.2026).

Сам рендер файлов проверяется в tests/test_analytics_export.py; здесь — поведение роутера,
которое из чистых функций не видно:
  • «скачать весь отчёт» (`report=all`) не должен разваливаться из-за одного тяжёлого раздела —
    заказчик получал пустой 500 вместо файла, где проблемный лист помечен «нет данных»;
  • одиночный отчёт, наоборот, обязан честно падать: пустой файл там хуже ошибки;
  • нехватка шрифта — это 503 с текстом «что доложить», а не 500 (п.3).

БД не нужна: сессия подменяется заглушкой, а `crud.build_report` — фейком.
Запуск: python -m pytest tests/test_analytics_export_route.py
"""
from __future__ import annotations

import logging
import os
import sys
from io import BytesIO

import pytest

sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

openpyxl = pytest.importorskip("openpyxl")
from openpyxl import load_workbook  # noqa: E402

from fastapi.testclient import TestClient  # noqa: E402

from app import crud, schemas as sch  # noqa: E402
from app.config import settings  # noqa: E402
from app.db import get_session  # noqa: E402
from app.main import app  # noqa: E402
from app.services import analytics_export as ax  # noqa: E402

_MODELS = {
    "overview": sch.AnalyticsOverview,
    "questions": sch.AnalyticsQuestions,
    "unanswered": sch.AnalyticsUnanswered,
    "engagement": sch.AnalyticsEngagement,
    "exhibits": sch.AnalyticsExhibits,
    "routes": sch.AnalyticsRoutes,
    "recognition": sch.AnalyticsRecognition,
}


async def _no_session():
    """Ручка экспорта не трогает сессию сама — всё, что нужно, отдаёт фейковый build_report."""
    yield None


def _client(broken: str = ""):
    """TestClient с подменённым build_report; `broken` — имя отчёта, который падает."""
    async def build_report(session, name, dfrom, dto, force=False):
        if name == broken:
            raise RuntimeError("тяжёлый раздел не посчитался")
        return _MODELS[name]()

    app.dependency_overrides[get_session] = _no_session
    saved = crud.build_report
    crud.build_report = build_report
    client = TestClient(app, raise_server_exceptions=False)
    client.headers.update({"Authorization": f"Bearer {settings.admin_api_token}"})
    return client, (lambda: (setattr(crud, "build_report", saved), app.dependency_overrides.clear()))


def test_all_survives_a_failing_section():
    """Падение `routes` не должно уносить шесть уже посчитанных разделов (п.6)."""
    client, restore = _client(broken="routes")
    try:
        response = client.get("/admin/analytics/export?report=all&format=xlsx")
        assert response.status_code == 200
        book = load_workbook(BytesIO(response.content))
        assert ax._REPORT_TITLES["routes"] in book.sheetnames
        sheet = book[ax._REPORT_TITLES["routes"]]
        texts = [sheet.cell(row=r, column=1).value for r in range(1, sheet.max_row + 1)]
        assert "Нет данных за выбранный период" in texts
        # Остальные разделы на месте — ради них файл и качали.
        assert ax._REPORT_TITLES["overview"] in book.sheetnames
    finally:
        restore()


def test_single_report_still_fails_loudly():
    """Одиночный отчёт с пустыми данными выглядел бы как «аналитики нет» — честнее ошибка."""
    client, restore = _client(broken="routes")
    try:
        assert client.get("/admin/analytics/export?report=routes&format=xlsx").status_code == 500
        assert client.get("/admin/analytics/export?report=overview&format=xlsx").status_code == 200
    finally:
        restore()


def _forget_font() -> None:
    """Забыть зарегистрированный шрифт: reportlab кэширует его на весь процесс, и после
    любого удачного PDF в этой же сессии _register_fonts уже не заглядывает в файл."""
    try:
        from reportlab.pdfbase import pdfmetrics
    except ImportError:                                # pragma: no cover — зависит от окружения
        return
    pdfmetrics._fonts.pop(ax._FONT_NAME, None)
    pdfmetrics._fonts.pop(ax._FONT_BOLD_NAME, None)


def test_missing_font_is_503_with_instructions():
    """П.3: без шрифта — 503 и текст «что доложить в сборку», а не 500 без тела."""
    client, restore = _client()
    saved = (ax._FONT_CANDIDATES, ax._bundled_font_paths, ax.settings.analytics_pdf_font_path)
    ax._FONT_CANDIDATES, ax._bundled_font_paths = (), (lambda: [])
    ax.settings.analytics_pdf_font_path = None
    _forget_font()
    try:
        response = client.get("/admin/analytics/export?report=overview&format=pdf")
        assert response.status_code == 503
        assert "ANALYTICS_PDF_FONT_PATH" in response.json()["detail"]
    finally:
        ax._FONT_CANDIDATES, ax._bundled_font_paths, ax.settings.analytics_pdf_font_path = saved
        _forget_font()
        restore()


def test_file_name_of_the_full_report():
    """Имя файла общего отчёта — faberge-analytics-<from>-<to>.<ext> (п.6)."""
    client, restore = _client()
    try:
        response = client.get(
            "/admin/analytics/export?report=all&format=xlsx&from=2026-07-01&to=2026-08-06"
        )
        assert "faberge-analytics-2026-07-01-2026-08-06.xlsx" in response.headers["content-disposition"]
    finally:
        restore()


if __name__ == "__main__":
    logging.disable(logging.CRITICAL)          # тест «падающего раздела» пишет traceback в лог
    failures = 0
    for name, func in sorted(globals().items()):
        if name.startswith("test_") and callable(func):
            try:
                func()
                print(f"ok   {name}")
            except AssertionError as exc:
                failures += 1
                print(f"FAIL {name}: {exc}")
    print("—" * 40)
    print("все тесты пройдены" if not failures else f"провалено: {failures}")
    sys.exit(1 if failures else 0)
