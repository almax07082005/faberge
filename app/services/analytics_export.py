"""Выгрузка аналитических отчётов в .xlsx и .pdf (§11 ТЗ 03.08.2026, п.5–6 баг-репорта 06.08.2026).

В ТЗ явно указаны оба формата: музей сводит цифры в Excel сам, а PDF показывает
на совещании. Отчёт раскладывается в набор секций «заголовок + таблица»
(`Section`), из которых одинаково собираются оба файла — логика раскладки одна,
меняется только рендер.

Баг-репорт 06.08.2026:
* п.5 «сейчас выглядит некорректно» — даты приезжали строкой с микросекундами и
  таймзоной, ширины не подогнаны, заголовки не выделены. Файл формирует бэк, фронт
  только отдаёт его пользователю, поэтому вся вёрстка листа — здесь.
* п.6 «весь отчёт одним файлом» — `to_xlsx_all` / `to_pdf_all`: лист (раздел) на отчёт.

ТИПЫ ЯЧЕЕК. Заказчик сводит выгрузку формулами, поэтому дата обязана быть датой,
счётчик — целым, доля — процентом. Тип протаскивается ПО ЯЧЕЙКЕ (обёртка `Cell`),
а не по колонке: в таблице «Показатели» одна и та же колонка «Значение» держит и
счётчики (`total_visits`), и доли (`chat_conversion_rate`), и средние — колоночный
`formats: list[str]` эту таблицу описать не может. Обёртка одинаково читается обоими
рендерами: xlsx берёт из неё `number_format`, pdf — текстовое представление
(`_cell_text`), поэтому число в PDF и в Excel выглядит одинаково.

ТЕКСТ ИЗ ВНЕШНЕГО МИРА. В ячейки попадают реплики посетителей, и обе библиотеки
на них падают целым файлом, а не одной ячейкой: openpyxl — на управляющих
символах и строке длиннее 32 767 знаков, reportlab — на «<мастеру>» внутри
вопроса (разбирает абзац как мини-XML). Поэтому строки чистятся в `_as_cell`
(`_plain`) и экранируются на входе в PDF (`_pdf_text`).

Зависимости (`openpyxl`, `reportlab`) импортируются ЛЕНИВО, внутри функций:
модуль тянется только при обращении к /admin/analytics/export, а холодный старт
функции Yandex Cloud и остальной API от них не зависят.

Кириллица в PDF. Стандартные шрифты ReportLab (Helvetica) кириллицу не
содержат — без TTF-шрифта выходит лист с квадратами. Шрифт ищется по
`ANALYTICS_PDF_FONT_PATH` и стандартным путям систем (DejaVu на Linux, Arial
Unicode на macOS); если не найден — поднимаем `ExportError` с инструкцией, а не
отдаём заведомо нечитаемый файл. Для /health есть `font_available()` — тот же
поиск, но без исключений (п.3 баг-репорта: проверять шрифт ДО жалобы заказчика).
"""
from __future__ import annotations

import os
import re
from dataclasses import dataclass, field
from datetime import date, datetime, timezone
from io import BytesIO
from typing import Any, Dict, List, Optional, Sequence, Set, Tuple

from ..config import settings


class ExportError(Exception):
    """Выгрузка невозможна (нет библиотеки или шрифта). Роутер отдаёт 503."""

    def __init__(self, message: str) -> None:
        super().__init__(message)
        self.message = message


# ── Типизированная ячейка ────────────────────────────────────────────────────
# Допустимые типы. "float" — среднее/длительность, "percent" — доля 0..1
# (crud._rate округляет до 4 знаков, в Excel это ровно процентный формат).
FORMATS = ("text", "int", "float", "percent", "date", "datetime")


@dataclass
class Cell:
    """Значение с типом отображения. Кладётся в строку секции вместо голого значения."""

    value: Any
    fmt: str = "text"


@dataclass
class Section:
    """Одна таблица внутри отчёта: заголовок, шапка и строки.

    В строках лежат либо `Cell`, либо голые значения — тогда тип определяется
    автоматически (`_as_cell`): int → счётчик, float → дробное, строка вида
    `2026-08-06` / `2026-08-06T04:00:00.123456+00:00` → дата/таймстемп.
    """

    title: str
    headers: List[str]
    rows: List[List[Any]] = field(default_factory=list)


# Шесть отчётов дашборда — состав одиночной выгрузки (контракт роутера не менялся).
REPORTS = ("overview", "questions", "unanswered", "exhibits", "routes", "recognition")

# Состав режима «весь отчёт» (п.6). Кроме шести плиток дашборда сюда добавлена
# `engagement`: именно про её метрики — «Средний визит», «Конверсия в диалог»,
# «Глубина визита» — заказчик спрашивает в п.7, и в «полном отчёте» их отсутствие
# выглядело бы дырой.
ALL_REPORTS = REPORTS + ("engagement",)

# Имя, под которым роутер запрашивает режим «всё сразу» (?report=all).
ALL_REPORT_KEY = "all"

_REPORT_TITLES = {
    "overview": "Сводная аналитика",
    "questions": "Вопросы посетителей",
    "unanswered": "Вопросы без ответа гида",
    "exhibits": "Статистика по экспонатам",
    "routes": "Маршруты по залам",
    "recognition": "Качество распознавания",
    "engagement": "Метрики визита",
}

_ALL_TITLE = "Аналитика посетителей"

# Человеческие подписи скалярных метрик. Ключи, которых здесь нет, в таблицу
# «Показатели» не попадают (служебные from/to/updated_at выводятся в шапке).
_METRIC_LABELS: Dict[str, str] = {
    "total_sessions": "Сессий",
    "total_visits": "Визитов",
    "total_app_opens": "Открытий приложения",
    "total_recognitions": "Распознаваний",
    "recognition_success_rate": "Доля успешных распознаваний",
    "total_chat_messages": "Вопросов гиду",
    "total_audio_plays": "Прослушиваний озвучки",
    "total_questions": "Реплик посетителей",
    "unique_questions": "Различных формулировок",
    "total_clusters": "Смысловых групп вопросов",
    "total_unanswered": "Вопросов без ответа",
    "total_answered": "Вопросов с ответом",
    "unanswered_rate": "Доля вопросов без ответа",
    "unclassified": "Без признака (до 03.08.2026)",
    "avg_duration_sec": "Средняя длительность визита, с",
    "median_duration_sec": "Медианная длительность визита, с",
    "max_duration_sec": "Максимальная длительность визита, с",
    "avg_events_per_session": "Событий за визит",
    "avg_exhibits_per_session": "Экспонатов за визит",
    "avg_questions_per_session": "Вопросов за визит",
    "sessions_with_chat": "Визитов с открытием чата",
    "sessions_with_questions": "Визитов с вопросом гиду",
    # Знаменатель конверсий (п.7 баг-репорта) — музей должен видеть его рядом с
    # самими конверсиями, иначе «12 %» не от чего отсчитать. Раньше на лист уезжали
    # только «Визитов», «Визитов с открытием чата» и сама доля — и ни одно из
    # частных не сходилось с процентом, потому что числитель считается ВНУТРИ базы
    # (см. schemas.AnalyticsEngagement): 20/96 и 20/141 против напечатанных 12,5 %.
    # Поэтому выгружаем всю дробь: база, знаменатель, числитель, доля — подряд.
    "sessions_with_app_open": "Визитов с открытием приложения",
    "conversion_basis": "База конверсий",
    "conversion_denominator": "Знаменатель конверсий (визитов)",
    "chat_conversion_numerator": "Из них открыли чат",
    "chat_conversion_rate": "Конверсия в чат",
    "question_conversion_numerator": "Из них задали вопрос",
    "question_conversion_rate": "Конверсия в вопрос",
    "total_sessions_with_route": "Визитов с маршрутом",
    "avg_halls_per_session": "Залов за визит",
    "total_devices": "Устройств",
    "returning_devices": "Вернувшихся устройств",
    "avg_sessions_per_device": "Сессий на устройство",
    "total_exhibits": "Экспонатов в каталоге",
    "never_viewed": "Ни разу не открывали",
    "total": "Попыток распознавания",
    "success": "Успешных",
    "success_rate": "Доля успешных",
    "fallback_shown": "Показан топ-3",
    "fallback_rate": "Доля показов топ-3",
    "fallback_converted": "Открыли карточку из топ-3",
    "fallback_conversion_rate": "Конверсия топ-3",
    "failed": "Неуспешных",
    "abandoned_after_fail": "Ушли после неудачи",
    "abandonment_rate": "Доля ушедших",
    "retry_after_fail": "Повторили съёмку",
    "avg_confidence": "Средняя уверенность",
}

# Доли 0..1 без суффикса `_rate` — правило «ключ кончается на _rate» их не ловит.
_SHARE_KEYS = frozenset({"avg_confidence"})

# Расшифровка значений-перечислений. `conversion_basis` приезжает как `app_open` /
# `all_visits`, и в отчёте для музея такая строка ничего не значит — а по ней
# решается, сопоставимы ли доли двух периодов (schemas.AnalyticsEngagement прямо
# предупреждает: сравнивать можно только при одинаковой базе), п.7 баг-репорта.
_ENUM_LABELS: Dict[str, Dict[Any, str]] = {
    "conversion_basis": {
        "app_open": "визиты с запуском приложения (app_open)",
        "all_visits": "все визиты (app_open за период не приходил)",
    },
}

# Причины, по которым гид не ответил (payload.fail_reasons отчёта unanswered).
_FAIL_REASON_LABELS = {
    "no_context": "Не привязан экспонат (нет контекста)",
    "llm_refusal": "Гид отказался отвечать",
    "not_found": "Экспонат не найден",
    "error": "Ошибка обращения к модели",
}


def _metric_format(key: str, value: Any) -> str:
    """Тип метрики по её ключу: доли в payload — это float 0..1 с суффиксом `_rate`."""
    if isinstance(value, bool):
        return "text"
    if key.endswith("_rate") or key in _SHARE_KEYS:
        return "percent"
    if isinstance(value, int):
        return "int"
    if isinstance(value, float):
        return "float"
    return "text"


# ── Распознавание ISO-дат ────────────────────────────────────────────────────
# Шаблоны намеренно строгие: в ячейках лежат вопросы посетителей и названия
# экспонатов, и превращать в дату случайную строку с цифрами нельзя. Дата
# признаётся датой, только если строка целиком совпадает с шаблоном И
# разбирается `datetime.fromisoformat`.
_ISO_DATE_RE = re.compile(r"^\d{4}-\d{2}-\d{2}$")
_ISO_DATETIME_RE = re.compile(
    r"^\d{4}-\d{2}-\d{2}[T ]\d{2}:\d{2}(:\d{2}(\.\d+)?)?(Z|[+-]\d{2}:?\d{2})?$"
)


def _parse_iso(text: str) -> Optional[date]:
    """ISO-дата/таймстемп → date/datetime, всё остальное → None."""
    if _ISO_DATE_RE.match(text):
        try:
            return date.fromisoformat(text)
        except ValueError:
            return None
    if _ISO_DATETIME_RE.match(text):
        # Python < 3.11 не понимает суффикс 'Z' в fromisoformat — нормализуем сами,
        # чтобы выгрузка не зависела от версии рантайма (в образе 3.12, локально 3.14).
        try:
            return datetime.fromisoformat(text.replace("Z", "+00:00"))
        except ValueError:
            return None
    return None


# ── Санитайзер текста ────────────────────────────────────────────────────────
# В ячейках лежат реплики посетителей — то есть текст, пришедший из внешнего
# мира, а не из каталога. Два способа уронить им выгрузку проверены руками:
# * openpyxl падает `IllegalCharacterError` на управляющих символах (набор ровно
#   тот же, что в openpyxl.cell.cell.ILLEGAL_CHARACTERS_RE; \t \n \r законны);
# * Excel не принимает строку длиннее 32 767 символов, а «Другие формулировки»
#   склеивают варианты кластера, которых crud набирает до 5000 штук.
# Роняется при этом ВЕСЬ файл, а не одна ячейка, — поэтому чистим на входе.
_CONTROL_CHARS_RE = re.compile(r"[\000-\010\013\014\016-\037]")
_MAX_TEXT_LEN = 32_000


def _plain(text: str) -> str:
    """Строка, безопасная для ячейки Excel: без управляющих символов и не длиннее лимита."""
    cleaned = _CONTROL_CHARS_RE.sub("", text)
    if len(cleaned) > _MAX_TEXT_LEN:
        return cleaned[: _MAX_TEXT_LEN - 1] + "…"
    return cleaned


def _naive_utc(value: datetime) -> datetime:
    """Excel не умеет хранить таймзону в ячейке — приводим к UTC и снимаем tzinfo.

    К местному времени музея приводить нельзя: смещения в payload нет, а тихая
    подмена «+03:00» испортила бы сведение отчётов. Поэтому UTC + подпись в шапке.

    Микросекунды срезаются: заказчик жаловался ровно на них (п.5), а в ячейке Excel
    время всё равно хранится дробью от суток — доли секунды дают только мусор при
    любом развороте формата.
    """
    if value.tzinfo is None:
        return value.replace(microsecond=0)
    return value.astimezone(timezone.utc).replace(tzinfo=None, microsecond=0)


def _as_cell(value: Any) -> Cell:
    """Голое значение → типизированная ячейка (порядок проверок важен)."""
    if isinstance(value, Cell):
        # Готовую ячейку (их собирает `_scalars`) пропускаем через ту же чистку:
        # funnel для обоих рендеров один, и мусор из payload не должен его обойти.
        return Cell(_plain(value.value), value.fmt) if isinstance(value.value, str) else value
    if value is None:
        return Cell(None, "text")
    if isinstance(value, bool):
        return Cell("да" if value else "нет", "text")
    if isinstance(value, int):
        return Cell(value, "int")
    if isinstance(value, float):
        return Cell(value, "float")
    if isinstance(value, datetime):          # datetime — подкласс date, проверяем первым
        return Cell(_naive_utc(value), "datetime")
    if isinstance(value, date):
        return Cell(value, "date")
    if isinstance(value, str):
        parsed = _parse_iso(value)
        if isinstance(parsed, datetime):
            return Cell(_naive_utc(parsed), "datetime")
        if isinstance(parsed, date):
            return Cell(parsed, "date")
        return Cell(_plain(value), "text")
    return Cell(_plain(str(value)), "text")


def _cell_text(cell: Cell) -> str:
    """Текстовое представление ячейки — для PDF и для расчёта ширины колонки.

    Разделитель дробной части — запятая: файл уходит заказчику как есть, а в
    русском Excel/Numbers точка в PDF-отчёте рядом с запятой в xlsx выглядит браком.
    """
    value = cell.value
    if value is None or value == "":
        return ""
    if cell.fmt == "percent":
        return f"{value * 100:.1f}".replace(".", ",") + " %"
    if cell.fmt == "int":
        return str(value)
    if cell.fmt == "float":
        return f"{value:.2f}".rstrip("0").rstrip(".").replace(".", ",") or "0"
    if cell.fmt == "date":
        return value.strftime("%d.%m.%Y")
    if cell.fmt == "datetime":
        return value.strftime("%d.%m.%Y %H:%M")
    return str(value)


# Числовые типы: выравниваются вправо и в xlsx, и в pdf.
_NUMERIC_FORMATS = frozenset({"int", "float", "percent", "date", "datetime"})

_NUMBER_FORMATS = {
    "int": "0",
    "float": "0.##",
    "percent": "0.0%",
    "date": "DD.MM.YYYY",
    "datetime": "DD.MM.YYYY HH:MM",
}


# ── Раскладка отчёта по секциям ──────────────────────────────────────────────
def _scalars(payload: Dict[str, Any]) -> Section:
    rows: List[List[Any]] = []
    for key, label in _METRIC_LABELS.items():
        if key not in payload:
            continue
        value = payload[key]
        if isinstance(value, (list, dict)):
            continue
        decoded = _ENUM_LABELS.get(key, {}).get(value)
        if decoded is not None:
            rows.append([label, Cell(decoded, "text")])
            continue
        rows.append([label, Cell(value, _metric_format(key, value))])
    return Section("Показатели", ["Показатель", "Значение"], rows)


def _top_section(title: str, items: Sequence[dict], name_header: str = "Название") -> Section:
    return Section(
        title,
        ["ID", name_header, "Количество"],
        [[item.get("id"), item.get("name"), item.get("count")] for item in items],
    )


def build_sections(report: str, payload: Dict[str, Any]) -> List[Section]:
    """Разложить JSON отчёта в таблицы для выгрузки."""
    sections: List[Section] = [_scalars(payload)]

    if report == "overview":
        sections.append(_top_section("Топ экспонатов", payload.get("top_exhibits", []), "Экспонат"))
        sections.append(_top_section("Топ залов", payload.get("top_halls", []), "Зал"))

    elif report == "questions":
        for key, title in (("frequent", "Частые вопросы"), ("rare", "Редкие вопросы")):
            sections.append(
                Section(
                    title,
                    ["Вопрос", "Количество", "Другие формулировки"],
                    [
                        [item.get("question"), item.get("count"), "; ".join(item.get("variants") or [])]
                        for item in payload.get(key, [])
                    ],
                )
            )

    elif report == "unanswered":
        # Сводка по причинам: музею она нужнее списка вопросов — показывает, чинить
        # описания экспонатов или промпт гида.
        reasons = payload.get("fail_reasons") or {}
        sections.append(
            Section(
                "Причины отсутствия ответа",
                ["Причина", "Количество"],
                [
                    [_FAIL_REASON_LABELS.get(key, key), count]
                    for key, count in sorted(reasons.items(), key=lambda kv: -kv[1])
                ],
            )
        )
        sections.append(
            Section(
                "Вопросы без ответа",
                ["Вопрос", "Количество", "Причины", "Экспонаты", "Другие формулировки"],
                [
                    [
                        item.get("question"),
                        item.get("count"),
                        ", ".join(
                            f"{_FAIL_REASON_LABELS.get(k, k)}: {v}"
                            for k, v in (item.get("fail_reasons") or {}).items()
                        ),
                        ", ".join(
                            str(e.get("name") or e.get("id")) for e in (item.get("exhibits") or [])
                        ),
                        "; ".join(item.get("variants") or []),
                    ]
                    for item in payload.get("items", [])
                ],
            )
        )

    elif report == "exhibits":
        sections.append(
            Section(
                "Экспонаты",
                ["ID", "Название", "Зал", "Просмотры", "Вопросы", "Озвучки", "Распознавания"],
                [
                    [
                        item.get("id"), item.get("name"), item.get("hall_number"),
                        item.get("views"), item.get("questions"),
                        item.get("tts_plays"), item.get("recognitions"),
                    ]
                    for item in payload.get("items", [])
                ],
            )
        )

    elif report == "engagement":
        # Гистограмма длительностей — то, чем заказчик проверяет «Средний визит» (п.7).
        sections.append(
            Section(
                "Длительность визита",
                ["Интервал", "Визитов"],
                [[item.get("label"), item.get("count")] for item in payload.get("buckets", [])],
            )
        )

    elif report == "routes":
        sections.append(_top_section("Посещения залов", payload.get("top_hall_visits", []), "Зал"))
        sections.append(_top_section("Залы входа", payload.get("top_entry_halls", []), "Зал"))
        sections.append(_top_section("Залы выхода", payload.get("top_exit_halls", []), "Зал"))
        sections.append(
            Section(
                "Экраны выхода",
                ["Экран", "Количество"],
                [[item.get("name"), item.get("count")] for item in payload.get("top_exit_screens", [])],
            )
        )
        sections.append(
            Section(
                "Переходы между залами",
                ["Из зала", "В зал", "Количество"],
                [
                    [
                        item.get("from_hall_name") or item.get("from_hall_id"),
                        item.get("to_hall_name") or item.get("to_hall_id"),
                        item.get("count"),
                    ]
                    for item in payload.get("top_transitions", [])
                ],
            )
        )
        sections.append(
            Section(
                "Частые маршруты",
                ["Маршрут", "Количество"],
                [
                    [
                        " → ".join(str(h.get("name") or h.get("id")) for h in item.get("halls", [])),
                        item.get("count"),
                    ]
                    for item in payload.get("top_paths", [])
                ],
            )
        )
        sections.append(
            Section(
                "Визитов на устройство",
                ["Группа", "Устройств"],
                [
                    [item.get("label"), item.get("devices")]
                    for item in payload.get("sessions_per_device_hist", [])
                ],
            )
        )

    return [s for s in sections if s.rows]


def _period_label(payload: Dict[str, Any]) -> str:
    start, end = payload.get("from") or "начало", payload.get("to") or "сегодня"
    start_cell, end_cell = _as_cell(start), _as_cell(end)
    return f"Период: {_cell_text(start_cell)} — {_cell_text(end_cell)}"


def _report_title(report: str) -> str:
    return _REPORT_TITLES.get(report, report)


def file_name(report: str, dfrom: Optional[date], dto: Optional[date], extension: str) -> str:
    """Имя файла выгрузки. Для режима «всё сразу» — `faberge-analytics-…` (п.6)."""
    slug = "analytics" if report == ALL_REPORT_KEY else report
    return f"faberge-{slug}-{dfrom or 'all'}-{dto or 'all'}.{extension}"


def _generated_at() -> datetime:
    """Момент формирования файла. Без tz — в ячейку кладём UTC (см. `_naive_utc`)."""
    return datetime.now(timezone.utc).replace(tzinfo=None, microsecond=0)


def _merged_meta(payloads: Dict[str, Dict[str, Any]]) -> Dict[str, Any]:
    """Период и «данные на» для общего отчёта.

    Отчёты считаются одним запросом за один период, но кэш мог обновиться в разное
    время — берём САМЫЙ СВЕЖИЙ `updated_at`, иначе шапка обещала бы данные новее,
    чем в части разделов.
    """
    meta: Dict[str, Any] = {"from": None, "to": None, "updated_at": None}
    newest: Optional[datetime] = None
    for payload in payloads.values():
        if not isinstance(payload, dict):
            continue
        meta["from"] = meta["from"] or payload.get("from")
        meta["to"] = meta["to"] or payload.get("to")
        raw = payload.get("updated_at")
        parsed = _parse_iso(raw) if isinstance(raw, str) else (raw if isinstance(raw, datetime) else None)
        if isinstance(parsed, datetime):
            parsed = _naive_utc(parsed)
            if newest is None or parsed > newest:
                newest, meta["updated_at"] = parsed, parsed
    return meta


# ── XLSX ─────────────────────────────────────────────────────────────────────
# Потолок ширины колонки. Вопросы посетителей бывают на 300 символов — колонка на
# такой текст растянулась бы на два экрана, поэтому ширина режется, а текст
# переносится (`wrap_text`); высоту строки Excel и Numbers подберут сами.
_MAX_COL_WIDTH = 55
_MIN_COL_WIDTH = 9

_HEADER_FILL_COLOR = "DCE6F1"
_GRID_COLOR = "B0B0B0"


def _track_width(widths: Dict[int, int], column: int, text: str) -> None:
    longest = max((len(part) for part in text.split("\n")), default=0)
    width = min(max(longest + 2, _MIN_COL_WIDTH), _MAX_COL_WIDTH)
    widths[column] = max(widths.get(column, 0), width)


def _sheet_name(title: str, used: Set[str]) -> str:
    """Имя листа под ограничения Excel: ≤31 символа, без []:*?/\\, уникальное.

    Библиотека молча обрезала бы длинное имя и упала бы на дубле, а заказчику
    в общем отчёте (п.6) нужны говорящие вкладки — поэтому чистим сами.
    """
    name = re.sub(r"[\[\]:*?/\\]", " ", title)
    name = re.sub(r"\s+", " ", name).strip().strip("'")
    name = (name[:31].strip()) or "Отчёт"
    base, suffix = name, 2
    while name.casefold() in used:
        tail = f" {suffix}"
        name = f"{base[:31 - len(tail)].strip()}{tail}"
        suffix += 1
    used.add(name.casefold())
    return name


def _write_sheet(sheet, report: str, payload: Dict[str, Any], generated: datetime) -> int:
    """Отрисовать один отчёт на листе; вернуть число строк данных (0 — нет данных).

    Порядок ровно как просил заказчик (п.5): шапка с названием, периодом и датой
    формирования → секции таблицами → автоширина по содержимому.
    """
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

    title_font = Font(bold=True, size=14)
    label_font = Font(bold=True)
    head_font = Font(bold=True)
    head_fill = PatternFill("solid", fgColor=_HEADER_FILL_COLOR)
    thin = Side(style="thin", color=_GRID_COLOR)
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    widths: Dict[int, int] = {}

    def put(row: int, column: int, raw: Any, *, font=None, fill=None, wrap=False, boxed=False) -> None:
        cell = _as_cell(raw)
        target = sheet.cell(row=row, column=column)
        target.value = cell.value
        number_format = _NUMBER_FORMATS.get(cell.fmt)
        if number_format:
            target.number_format = number_format
        if font is not None:
            target.font = font
        if fill is not None:
            target.fill = fill
        target.alignment = Alignment(
            # Числа/даты вправо, текст влево, всё — по верхнему краю: со включённым
            # переносом длинный вопрос иначе «плавает» по центру высокой строки.
            horizontal="right" if cell.fmt in _NUMERIC_FORMATS else "left",
            vertical="top",
            wrap_text=wrap,
        )
        if boxed:
            target.border = border
        _track_width(widths, column, _cell_text(cell))

    row = 1
    put(row, 1, _report_title(report), font=title_font)
    row += 1
    put(row, 1, "Период", font=label_font)
    put(row, 2, payload.get("from") or "начало")
    put(row, 3, payload.get("to") or "сегодня")
    row += 1
    put(row, 1, "Данные на", font=label_font)
    put(row, 2, payload.get("updated_at") or "—")
    put(row, 3, "UTC")
    row += 1
    put(row, 1, "Файл создан", font=label_font)
    put(row, 2, generated)
    put(row, 3, "UTC")
    row += 2

    sections = build_sections(report, payload)
    if not sections:
        put(row, 1, "Нет данных за выбранный период")
        _apply_widths(sheet, widths)
        return 0

    freeze_row: Optional[int] = None
    total_rows = 0
    for section in sections:
        put(row, 1, section.title, font=label_font)
        row += 1
        for column, header in enumerate(section.headers, start=1):
            put(row, column, header, font=head_font, fill=head_fill, wrap=True, boxed=True)
        if freeze_row is None:
            # Закрепляем по ПЕРВОЙ секции: на листе их несколько, и закрепить шапку
            # каждой нельзя — freeze_panes в Excel один на лист. Первая секция
            # («Показатели») плюс шапка отчёта отвечают на главный вопрос при
            # прокрутке — «что это за цифры и за какой период».
            freeze_row = row + 1
        row += 1
        for data_row in section.rows:
            for column, value in enumerate(data_row, start=1):
                put(row, column, value, wrap=True, boxed=True)
            row += 1
            total_rows += 1
        row += 1

    if freeze_row:
        sheet.freeze_panes = f"A{freeze_row}"
    _apply_widths(sheet, widths)
    return total_rows


def _apply_widths(sheet, widths: Dict[int, int]) -> None:
    from openpyxl.utils import get_column_letter

    for column, width in widths.items():
        sheet.column_dimensions[get_column_letter(column)].width = width


def to_xlsx(report: str, payload: Dict[str, Any]) -> bytes:
    """Один лист на отчёт: шапка периода, секции таблицами, числа — числами."""
    try:
        from openpyxl import Workbook
    except ImportError as exc:  # pragma: no cover — зависит от окружения
        raise ExportError("Для выгрузки в .xlsx нужен пакет openpyxl.") from exc

    book = Workbook()
    sheet = book.active
    sheet.title = _sheet_name(_report_title(report), set())
    _write_sheet(sheet, report, payload, _generated_at())

    buffer = BytesIO()
    book.save(buffer)
    return buffer.getvalue()


def to_xlsx_all(payloads: Dict[str, Dict[str, Any]]) -> bytes:
    """Весь отчёт одним файлом: лист-оглавление + лист на каждый отчёт (п.6).

    Отчёт, которого нет в `payloads` (или пустой), лист всё равно получает — с
    пометкой «нет данных»: заказчик должен видеть, что раздел не потерялся.
    """
    try:
        from openpyxl import Workbook
    except ImportError as exc:  # pragma: no cover — зависит от окружения
        raise ExportError("Для выгрузки в .xlsx нужен пакет openpyxl.") from exc

    generated = _generated_at()
    meta = _merged_meta(payloads)

    book = Workbook()
    index = book.active
    used: Set[str] = set()
    index.title = _sheet_name("Оглавление", used)

    contents: List[Tuple[str, str, int]] = []
    for report in ALL_REPORTS:
        payload = payloads.get(report) or {}
        sheet = book.create_sheet(_sheet_name(_report_title(report), used))
        rows = _write_sheet(sheet, report, payload, generated)
        contents.append((_report_title(report), sheet.title, rows))

    _write_index(index, meta, generated, contents)

    buffer = BytesIO()
    book.save(buffer)
    return buffer.getvalue()


def _write_index(sheet, meta: Dict[str, Any], generated: datetime, contents) -> None:
    """Первый лист общего файла: что внутри и за какой период."""
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

    head_fill = PatternFill("solid", fgColor=_HEADER_FILL_COLOR)
    thin = Side(style="thin", color=_GRID_COLOR)
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    widths: Dict[int, int] = {}

    def put(row: int, column: int, raw: Any, *, font=None, fill=None, boxed=False) -> None:
        cell = _as_cell(raw)
        target = sheet.cell(row=row, column=column)
        target.value = cell.value
        number_format = _NUMBER_FORMATS.get(cell.fmt)
        if number_format:
            target.number_format = number_format
        if font is not None:
            target.font = font
        if fill is not None:
            target.fill = fill
        target.alignment = Alignment(
            horizontal="right" if cell.fmt in _NUMERIC_FORMATS else "left", vertical="top"
        )
        if boxed:
            target.border = border
        _track_width(widths, column, _cell_text(cell))

    put(1, 1, _ALL_TITLE, font=Font(bold=True, size=14))
    put(2, 1, "Период", font=Font(bold=True))
    put(2, 2, meta.get("from") or "начало")
    put(2, 3, meta.get("to") or "сегодня")
    put(3, 1, "Данные на", font=Font(bold=True))
    put(3, 2, meta.get("updated_at") or "—")
    put(3, 3, "UTC")
    put(4, 1, "Файл создан", font=Font(bold=True))
    put(4, 2, generated)
    put(4, 3, "UTC")

    row = 6
    for column, header in enumerate(("Раздел", "Лист", "Строк"), start=1):
        put(row, column, header, font=Font(bold=True), fill=head_fill, boxed=True)
    row += 1
    for title, sheet_name, rows in contents:
        put(row, 1, title, boxed=True)
        put(row, 2, sheet_name, boxed=True)
        put(row, 3, rows if rows else "нет данных", boxed=True)
        row += 1

    sheet.freeze_panes = "A7"
    _apply_widths(sheet, widths)


# ── PDF ──────────────────────────────────────────────────────────────────────
# Шрифты с кириллицей: сначала явно заданный в конфигурации, затем каталог
# приложения (кладётся в образ), затем стандартные места установки в Linux-образах
# (fonts-dejavu-core / fonts-liberation / fonts-freefont в Debian и Ubuntu) и на macOS.
_FONT_CANDIDATES = (
    "/usr/share/fonts/truetype/dejavu/DejaVuSans.ttf",       # fonts-dejavu-core, Debian/Ubuntu
    "/usr/share/fonts/truetype/ttf-dejavu/DejaVuSans.ttf",   # старые образы Debian
    "/usr/share/fonts/dejavu/DejaVuSans.ttf",                # Fedora/RHEL
    "/usr/share/fonts/TTF/DejaVuSans.ttf",                   # Arch/Alpine
    "/usr/local/share/fonts/DejaVuSans.ttf",                 # ручная доустановка в образ
    "/usr/share/fonts/truetype/liberation/LiberationSans-Regular.ttf",
    "/usr/share/fonts/truetype/liberation2/LiberationSans-Regular.ttf",
    "/usr/share/fonts/truetype/freefont/FreeSans.ttf",       # fonts-freefont-ttf
    "/usr/share/fonts/truetype/noto/NotoSans-Regular.ttf",   # fonts-noto-core
    "/Library/Fonts/DejaVuSans.ttf",
    "/System/Library/Fonts/Supplemental/Arial Unicode.ttf",
    "/System/Library/Fonts/Supplemental/Arial.ttf",
    "/Library/Fonts/Arial Unicode.ttf",
)
_FONT_NAME = "AnalyticsCyrillic"
_FONT_BOLD_NAME = "AnalyticsCyrillic-Bold"


def _pdf_text(text: Any) -> str:
    """Текст для `Paragraph`: reportlab разбирает его как мини-XML, а не как строку.

    Вопрос посетителя «сколько лет <мастеру>?» превращается в незакрытый тег и
    роняет сборку целиком (`paraparser: syntax error`) — то есть ровно «PDF не
    формируется» из п.3 баг-репорта, только по вине данных, а не шрифта. Плюс
    вопрос с `<b>` молча приехал бы жирным. Экранируем всё, что уходит в абзац,
    включая статические подписи — так не нужно помнить, какая строка «своя».
    """
    return (
        str(text)
        .replace("&", "&amp;")
        .replace("<", "&lt;")
        .replace(">", "&gt;")
    )


# Потолок текста в ячейке PDF. Строку выше страницы reportlab разбить не умеет и
# валит сборку `LayoutError` — а «Другие формулировки» склеивают до нескольких
# тысяч вариантов кластера. PDF заказчик читает и печатает, полные тексты берутся
# из xlsx, поэтому здесь обрезаем, а не ломаемся.
_PDF_CELL_MAX_CHARS = 700


def _pdf_clip(text: str) -> str:
    return text if len(text) <= _PDF_CELL_MAX_CHARS else text[: _PDF_CELL_MAX_CHARS - 1] + "…"


def _bundled_font_paths() -> List[str]:
    """Шрифт, положенный рядом с кодом.

    Два варианта пути: относительно модуля (обычный деплой в контейнере) и
    относительно рабочего каталога — в Yandex Cloud Functions архив распаковывается
    в /function/code и cwd указывает именно туда, а __file__ может быть абсолютным
    путём во временный каталог сборки.
    """
    root = os.path.dirname(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
    return [
        os.path.join(root, "assets", "fonts", "DejaVuSans.ttf"),
        os.path.join(os.getcwd(), "assets", "fonts", "DejaVuSans.ttf"),
        "/function/code/assets/fonts/DejaVuSans.ttf",
    ]


# Сигнатуры sfnt, которые ReportLab умеет читать (тот же набор, что проверяет
# scripts/fetch_pdf_font.py при закачке). 'OTTO' — OpenType с CFF-контурами:
# файл валиден, но TTFont на нём падает, поэтому о нём говорим отдельной строкой.
_TTF_MAGIC = (b"\x00\x01\x00\x00", b"true", b"ttcf")
_OTTO_MAGIC = b"OTTO"


def _font_defect(path: str) -> Optional[str]:
    """Почему файл по этому пути не годится как шрифт, или None, если годится.

    П.3 баг-репорта: заказчик жаловался, что «PDF не формируется». Существование
    файла ничего не гарантирует — на месте шрифта регулярно оказывается сохранённая
    страница прокси, обрывок закачки или OpenType/CFF. Раньше такой файл проходил
    `os.path.isfile`, /health рапортовал `pdf_font: up`, а выгрузка падала внутри
    reportlab `TTFError` — то есть 500 без текста вместо честного 503 с подсказкой,
    что доложить. Проверяем сигнатуру ДО регистрации, как это делает
    scripts/fetch_pdf_font.py при закачке.
    """
    try:
        with open(path, "rb") as handle:
            head = handle.read(4)
    except OSError as exc:
        return f"файл не читается ({exc.strerror or exc})"
    if not head:
        return "файл пустой"
    if head == _OTTO_MAGIC:
        return "это OpenType/CFF (OTTO) — ReportLab такой не читает, нужен TrueType"
    if not head.startswith(_TTF_MAGIC):
        return f"не похож на TrueType, первые байты {head!r}"
    return None


def _resolve_font_path() -> str:
    candidates = [settings.analytics_pdf_font_path] if settings.analytics_pdf_font_path else []
    candidates += _bundled_font_paths()
    candidates += list(_FONT_CANDIDATES)
    # Негодные файлы не молчим, а копим: администратору важно узнать, что шрифт
    # ЛЕЖИТ, но битый, — иначе он будет искать несуществующую проблему с путём.
    rejected: List[str] = []
    for path in candidates:
        if not path or not os.path.isfile(path):
            continue
        defect = _font_defect(path)
        if defect is None:
            return path
        rejected.append(f"{path}: {defect}")
    detail = (" Найденные файлы не подошли: " + "; ".join(rejected) + ".") if rejected else ""
    raise ExportError(
        "Не найден шрифт с кириллицей для PDF. Укажите путь к TTF в переменной "
        "окружения ANALYTICS_PDF_FONT_PATH (например, DejaVuSans.ttf) или положите "
        "файл в assets/fonts/DejaVuSans.ttf. В Debian/Ubuntu-образе достаточно "
        "поставить пакет fonts-dejavu-core." + detail
    )


def font_available() -> bool:
    """Есть ли ПРИГОДНЫЙ шрифт для PDF. Для /health (п.3): проверка ничего не роняет.

    Именно пригодный: битый файл на месте шрифта раньше давал `pdf_font: up` и 500
    при нажатии кнопки — /health обязан гореть красным до жалобы заказчика, а не после.
    """
    try:
        return bool(_resolve_font_path())
    except ExportError:
        return False
    except Exception:  # pragma: no cover — перестраховка: /health не падает никогда
        return False


def _bold_candidates(regular: str) -> List[str]:
    """Жирное начертание рядом с обычным: DejaVuSans-Bold.ttf, «Arial Bold.ttf» и т.п."""
    base, ext = os.path.splitext(regular)
    return [f"{base}-Bold{ext}", f"{base} Bold{ext}", f"{base}Bold{ext}", f"{base}-Bd{ext}"]


def _register_fonts() -> Tuple[str, str]:
    """Зарегистрировать шрифт(ы) и вернуть (обычный, жирный).

    Жирный опционален: если рядом с TTF его нет (Arial Unicode на macOS), заголовки
    остаются обычным начертанием — это лучше, чем 503 из-за отсутствия второго файла.
    """
    from reportlab.pdfbase import pdfmetrics
    from reportlab.pdfbase.ttfonts import TTFont

    if _FONT_NAME not in pdfmetrics.getRegisteredFontNames():
        regular = _resolve_font_path()
        try:
            pdfmetrics.registerFont(TTFont(_FONT_NAME, regular))
        except Exception as exc:
            # Последний рубеж: сигнатуру файл прошёл, а внутри всё равно битый
            # (обрезанная закачка, экзотический вариант sfnt). Без этого наружу
            # улетал reportlab TTFError, роутер ловит только ExportError — и
            # заказчик получал 500 без текста вместо 503 с подсказкой (п.3).
            raise ExportError(
                f"Файл шрифта {regular} не читается ReportLab ({type(exc).__name__}: {exc}). "
                "Похоже, он повреждён или это не TrueType: перекачайте его "
                "(scripts/fetch_pdf_font.py) либо укажите другой в ANALYTICS_PDF_FONT_PATH."
            ) from exc
        for path in _bold_candidates(regular):
            if os.path.isfile(path):
                try:
                    pdfmetrics.registerFont(TTFont(_FONT_BOLD_NAME, path))
                except Exception:  # pragma: no cover — битый файл шрифта не ломает выгрузку
                    pass
                break
    bold = _FONT_BOLD_NAME if _FONT_BOLD_NAME in pdfmetrics.getRegisteredFontNames() else _FONT_NAME
    return _FONT_NAME, bold


def _pdf_styles() -> Dict[str, Any]:
    from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet

    regular, bold = _register_fonts()
    base = getSampleStyleSheet()
    return {
        "title": ParagraphStyle("t", parent=base["Title"], fontName=bold, fontSize=16),
        "meta": ParagraphStyle("m", parent=base["Normal"], fontName=regular, fontSize=9),
        "head": ParagraphStyle("h", parent=base["Heading2"], fontName=bold, fontSize=12),
        "section": ParagraphStyle("s", parent=base["Heading3"], fontName=bold, fontSize=11),
        "cell": ParagraphStyle("c", parent=base["Normal"], fontName=regular, fontSize=8, leading=10),
        "cell_right": ParagraphStyle(
            "cr", parent=base["Normal"], fontName=regular, fontSize=8, leading=10, alignment=2
        ),
        "cell_head": ParagraphStyle("ch", parent=base["Normal"], fontName=bold, fontSize=8, leading=10),
    }


def _column_widths(columns: List[List[str]], available: float) -> List[float]:
    """Ширины колонок пропорционально содержимому, в сумме ровно на ширину полосы.

    Без явных ширин reportlab растягивает таблицу по самой длинной строке и она
    уезжает за край листа — заказчик как раз жаловался, что таблица разъезжается.
    """
    weights = [
        max(6.0, min(60.0, float(max((len(text) for text in column), default=6))))
        for column in columns
    ]
    total = sum(weights) or 1.0
    return [available * weight / total for weight in weights]


def _section_flowables(section: Section, styles: Dict[str, Any], available: float) -> List[Any]:
    from reportlab.lib import colors
    from reportlab.lib.units import mm
    from reportlab.platypus import Paragraph, Spacer, Table, TableStyle

    cells = [[_as_cell(value) for value in row] for row in section.rows]
    texts = [[_pdf_clip(_cell_text(cell)) for cell in row] for row in cells]
    width = len(section.headers)
    columns = [
        [section.headers[index]] + [row[index] for row in texts if index < len(row)]
        for index in range(width)
    ]
    # Колонка считается числовой, если ВСЕ её непустые значения числовые —
    # тогда её содержимое прижимается вправо, как в xlsx.
    numeric = [
        all(
            row[index].fmt in _NUMERIC_FORMATS
            for row in cells
            if index < len(row) and row[index].value not in (None, "")
        )
        and any(index < len(row) for row in cells)
        for index in range(width)
    ]

    data = [[Paragraph(_pdf_text(header), styles["cell_head"]) for header in section.headers]]
    for row in texts:
        data.append(
            [
                Paragraph(_pdf_text(row[index]) if index < len(row) else "",
                          styles["cell_right"] if numeric[index] else styles["cell"])
                for index in range(width)
            ]
        )

    table = Table(data, colWidths=_column_widths(columns, available), repeatRows=1, hAlign="LEFT")
    table.setStyle(
        TableStyle(
            [
                ("GRID", (0, 0), (-1, -1), 0.3, colors.HexColor("#B0B0B0")),
                ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#DCE6F1")),
                ("VALIGN", (0, 0), (-1, -1), "TOP"),
                ("LEFTPADDING", (0, 0), (-1, -1), 4),
                ("RIGHTPADDING", (0, 0), (-1, -1), 4),
            ]
        )
    )
    return [Paragraph(_pdf_text(section.title), styles["section"]), table, Spacer(1, 5 * mm)]


def _report_flowables(
    report: str, payload: Dict[str, Any], styles: Dict[str, Any], available: float, *, with_meta: bool
) -> List[Any]:
    """Раздел одного отчёта: заголовок, при необходимости период, таблицы."""
    from reportlab.lib.units import mm
    from reportlab.platypus import Paragraph, Spacer

    story: List[Any] = [Paragraph(_pdf_text(_report_title(report)), styles["head"])]
    if with_meta:
        story.append(Paragraph(_pdf_text(_period_label(payload)), styles["meta"]))
        updated = _cell_text(_as_cell(payload.get("updated_at"))) or "—"
        story.append(Paragraph(_pdf_text(f"Данные на: {updated} UTC"), styles["meta"]))
    story.append(Spacer(1, 4 * mm))

    sections = build_sections(report, payload)
    if not sections:
        story.append(Paragraph("Нет данных за выбранный период.", styles["meta"]))
        return story
    for section in sections:
        story += _section_flowables(section, styles, available)
    return story


def _pdf_document(buffer: BytesIO, title: str):
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib.units import mm
    from reportlab.platypus import SimpleDocTemplate

    return SimpleDocTemplate(
        buffer, pagesize=landscape(A4),
        leftMargin=12 * mm, rightMargin=12 * mm, topMargin=12 * mm, bottomMargin=12 * mm,
        title=title,
    )


def to_pdf(report: str, payload: Dict[str, Any]) -> bytes:
    """Таблицы отчёта + шапка с периодом и датой формирования."""
    try:
        from reportlab.lib.units import mm
        from reportlab.platypus import Paragraph, Spacer
    except ImportError as exc:  # pragma: no cover — зависит от окружения
        raise ExportError("Для выгрузки в .pdf нужен пакет reportlab.") from exc

    styles = _pdf_styles()
    buffer = BytesIO()
    doc = _pdf_document(buffer, _report_title(report))
    generated = _generated_at().strftime("%d.%m.%Y %H:%M")
    updated = _cell_text(_as_cell(payload.get("updated_at"))) or "—"

    story: List[Any] = [
        Paragraph(_pdf_text(_report_title(report)), styles["title"]),
        Paragraph(_pdf_text(_period_label(payload)), styles["meta"]),
        Paragraph(_pdf_text(f"Данные на: {updated} UTC"), styles["meta"]),
        Paragraph(_pdf_text(f"Файл сформирован: {generated} UTC"), styles["meta"]),
        Spacer(1, 6 * mm),
    ]
    sections = build_sections(report, payload)
    if not sections:
        story.append(Paragraph("Нет данных за выбранный период.", styles["meta"]))
    for section in sections:
        story += _section_flowables(section, styles, doc.width)

    doc.build(story)
    return buffer.getvalue()


def to_pdf_all(payloads: Dict[str, Dict[str, Any]]) -> bytes:
    """Весь отчёт одним PDF: общая шапка периода, раздел на отчёт (п.6).

    Каждый раздел начинается с новой страницы: заказчик печатает такой отчёт и
    раздаёт по разделам, а раздел, начавшийся в середине предыдущей страницы,
    так не отделить. Титульная страница со списком разделов — как оглавление в xlsx.
    """
    try:
        from reportlab.lib.units import mm
        from reportlab.platypus import PageBreak, Paragraph, Spacer
    except ImportError as exc:  # pragma: no cover — зависит от окружения
        raise ExportError("Для выгрузки в .pdf нужен пакет reportlab.") from exc

    styles = _pdf_styles()
    meta = _merged_meta(payloads)
    generated = _generated_at().strftime("%d.%m.%Y %H:%M")
    updated = _cell_text(_as_cell(meta.get("updated_at"))) or "—"

    buffer = BytesIO()
    doc = _pdf_document(buffer, _ALL_TITLE)
    story: List[Any] = [
        Paragraph(_pdf_text(_ALL_TITLE), styles["title"]),
        Paragraph(_pdf_text(_period_label(meta)), styles["meta"]),
        Paragraph(_pdf_text(f"Данные на: {updated} UTC"), styles["meta"]),
        Paragraph(_pdf_text(f"Файл сформирован: {generated} UTC"), styles["meta"]),
        Spacer(1, 6 * mm),
        Paragraph("Разделы отчёта", styles["head"]),
    ]
    for number, report in enumerate(ALL_REPORTS, start=1):
        has_data = bool(build_sections(report, payloads.get(report) or {}))
        mark = "" if has_data else " — нет данных"
        story.append(Paragraph(_pdf_text(f"{number}. {_report_title(report)}{mark}"), styles["meta"]))

    for report in ALL_REPORTS:
        story.append(PageBreak())
        story += _report_flowables(
            report, payloads.get(report) or {}, styles, doc.width, with_meta=True
        )

    doc.build(story)
    return buffer.getvalue()
